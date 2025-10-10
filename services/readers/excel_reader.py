# services/readers/excel_reader.py

import io
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

logger = logging.getLogger("excel_reader")

def _clean_cell_value(value):
    """
    Converts cell value to a standardized string, handling None/NaN/dates.
    """
    if pd.isna(value) or value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()

def get_excel_content(file_obj, file_name):
    """
    Extracts text content from .xlsx and .xls files by iterating through worksheets.
    """
    full_text = []
    
    # Reset file pointer
    file_obj.seek(0)
    
    extension = file_name.lower().split('.')[-1]
    
    if extension == 'xlsx':
        # Use openpyxl for .xlsx
        try:
            # openpyxl requires an open file-like object
            wb = load_workbook(file_obj, read_only=True, data_only=True)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                full_text.append(f"\n===== Worksheet: {sheet_name} =====\n")
                
                for row in ws.rows:
                    row_data = [_clean_cell_value(cell.value) for cell in row]
                    # একটি সারির সব ডেটা ট্যাব দিয়ে যুক্ত করে
                    full_text.append('\t'.join(row_data)) 
                
        except InvalidFileException:
            logger.error(f"Invalid Excel file (xlsx): {file_name}")
            return None
        except Exception as e:
            logger.error(f"Error reading Excel file {file_name} with openpyxl: {e}")
            return None
            
    elif extension == 'xls':
        # Use pandas for .xls (requires xlrd backend, which should be included with pandas)
        try:
            # pandas.read_excel can handle the file-like object directly
            excel_data = pd.read_excel(file_obj, sheet_name=None, engine='xlrd')
            
            for sheet_name, df in excel_data.items():
                full_text.append(f"\n===== Worksheet: {sheet_name} =====\n")
                # প্রতিটি সারিকে স্ট্রিং এ রূপান্তর করে লিস্টে যোগ করা হলো
                for row_index, row in df.iterrows():
                    row_data = [_clean_cell_value(item) for item in row]
                    full_text.append('\t'.join(row_data))
                    
        except Exception as e:
            logger.error(f"Error reading Excel file {file_name} with pandas/xlrd: {e}")
            return None
            
    else:
        logger.warning(f"Unsupported extension for Excel reader: {extension}")
        return None

    # Line endings Standardization: This is crucial for search_text_content
    # The excel content is separated by newline characters (and tabs within a line)
    return '\n'.join(full_text)


def get_csv_content(file_obj):
    """
    Extracts text content from .csv files.
    """
    # Reset file pointer
    file_obj.seek(0)
    
    try:
        # Use pandas to read CSV robustly, including encoding detection (if possible)
        df = pd.read_csv(file_obj, encoding='utf-8', on_bad_lines='skip', sep=None, engine='python')
        
        # Convert DataFrame to a single string representation
        # Using to_csv ensures consistent delimiter and quote handling
        csv_string = df.to_csv(index=False, sep='\t', header=True)
        
        # Clean up line endings as done in text_reader for consistency
        return csv_string.replace('\r\n', '\n').replace('\r', '\n')
        
    except Exception as e:
        logger.error(f"Error reading CSV file with pandas: {e}")
        # Fallback: Try decoding as pure text (as done in text_reader)
        file_obj.seek(0)
        try:
             content_bytes = file_obj.read()
             return content_bytes.decode('utf-8').replace('\r\n', '\n').replace('\r', '\n')
        except:
             return None